from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from difflib import SequenceMatcher
from io import BytesIO
import re
from calendar import monthrange

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.models.inventory_snapshot import InventorySnapshot
from app.models.product import Product, ProductCodeHistory
from app.models.stock_movement import MovementType, StockMovement
from app.models.upload import UploadBatch, UploadStatus
from app.schemas.inventory import (
    InventoryListItem,
    InventoryListResponse,
    InventoryMonthlyItem,
    InventoryMonthlyResponse,
)
from app.schemas.upload import (
    UploadBatchItem,
    UploadBatchListResponse,
    UploadCommitRequest,
    UploadPreviewError,
    UploadPreviewProduct,
    UploadPreviewResponse,
    UploadPreviewRow,
    UploadPreviewSummary,
)


SOURCE_MONTHLY_INVENTORY = "monthly_inventory"

REQUIRED_COLUMNS = [
    "회사명",
    "상품코드",
    "공급사",
    "상품명",
    "바코드",
    "구분",
    "현재고",
    "금월 마감재고",
    "전월 마감재고",
    "마감재고 증감",
    "입고",
    "창고 내 입고",
    "반품 입고",
    "출고",
    "반출",
    "회송",
    "재고조정 입고",
    "재고조정 반출",
    "증감",
]

NUMERIC_FIELDS = {
    "현재고": "current_stock",
    "금월 마감재고": "closing_stock_current_month",
    "전월 마감재고": "closing_stock_previous_month",
    "마감재고 증감": "stock_change",
    "입고": "inbound_quantity",
    "창고 내 입고": "warehouse_inbound_quantity",
    "반품 입고": "return_inbound_quantity",
    "출고": "outbound_quantity",
    "반출": "carryout_quantity",
    "회송": "return_outbound_quantity",
    "재고조정 입고": "adjustment_inbound_quantity",
    "재고조정 반출": "adjustment_outbound_quantity",
    "증감": "net_change",
}


@dataclass
class ProductMatch:
    product: Product | None
    method: str | None


def parse_monthly_inventory_preview(
    *,
    db: Session,
    file_name: str,
    content: bytes,
    year: int | None = None,
    month: int | None = None,
    source: str = SOURCE_MONTHLY_INVENTORY,
) -> UploadPreviewResponse:
    target_year, target_month = _resolve_year_month(file_name, year, month)
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=False)
    sheet = workbook.active

    headers = _read_headers(sheet)
    missing_columns = [column for column in REQUIRED_COLUMNS if column not in headers]
    existing_committed_batches = _count_existing_batches(db, target_year, target_month, source)

    if missing_columns:
        return UploadPreviewResponse(
            summary=UploadPreviewSummary(
                file_name=file_name,
                year=target_year,
                month=target_month,
                source=source,
                total_rows=0,
                preview_rows=0,
                matched_rows=0,
                unmatched_rows=0,
                duplicate_rows=0,
                error_rows=1,
                missing_columns=missing_columns,
                existing_committed_batches=existing_committed_batches,
                should_confirm_replace=existing_committed_batches > 0,
            ),
            rows=[],
            errors=[
                UploadPreviewError(
                    row_number=1,
                    error_message=f"필수 컬럼이 없습니다: {', '.join(missing_columns)}",
                    raw_data={},
                )
            ],
        )

    products = list(db.scalars(select(Product)).all())
    seen_keys: set[str] = set()
    rows: list[UploadPreviewRow] = []
    errors: list[UploadPreviewError] = []

    for row_number, values in _iter_data_rows(sheet, headers):
        raw_data = {column: _clean_cell(values.get(column)) for column in REQUIRED_COLUMNS}
        try:
            product_code = _text(raw_data.get("상품코드"))
            barcode = _text(raw_data.get("바코드"))
            external_code, product_name = split_external_product_name(
                _text(raw_data.get("상품명")) or ""
            )
            if not product_name:
                raise ValueError("상품명이 비어 있습니다.")

            match = match_product(
                products=products,
                product_code=product_code,
                barcode=barcode,
                product_name=product_name,
            )
            duplicate_key = _duplicate_key(product_code, barcode, external_code, product_name)
            is_duplicate = duplicate_key in seen_keys
            seen_keys.add(duplicate_key)

            numeric_values = {
                field_name: _to_int(raw_data.get(column_name))
                for column_name, field_name in NUMERIC_FIELDS.items()
            }
            matched_product = None
            if match.product:
                matched_product = UploadPreviewProduct(
                    id=match.product.id,
                    product_code=match.product.product_code,
                    external_code=match.product.external_code,
                    barcode=match.product.barcode,
                    name=match.product.name,
                )

            rows.append(
                UploadPreviewRow(
                    row_number=row_number,
                    company_name=_text(raw_data.get("회사명")),
                    product_code=product_code,
                    supplier=_text(raw_data.get("공급사")),
                    external_code=external_code,
                    product_name=product_name,
                    barcode=barcode,
                    item_type=_text(raw_data.get("구분")),
                    matched=match.product is not None,
                    match_method=match.method,
                    matched_product=matched_product,
                    is_duplicate=is_duplicate,
                    duplicate_reason="업로드 파일 안에서 중복된 상품입니다." if is_duplicate else None,
                    raw_data=raw_data,
                    **numeric_values,
                )
            )
        except ValueError as exc:
            errors.append(
                UploadPreviewError(
                    row_number=row_number,
                    error_message=str(exc),
                    raw_data=raw_data,
                )
            )

    matched_rows = sum(1 for row in rows if row.matched)
    duplicate_rows = sum(1 for row in rows if row.is_duplicate)

    return UploadPreviewResponse(
        summary=UploadPreviewSummary(
            file_name=file_name,
            year=target_year,
            month=target_month,
            source=source,
            total_rows=len(rows) + len(errors),
            preview_rows=len(rows),
            matched_rows=matched_rows,
            unmatched_rows=len(rows) - matched_rows,
            duplicate_rows=duplicate_rows,
            error_rows=len(errors),
            missing_columns=[],
            existing_committed_batches=existing_committed_batches,
            should_confirm_replace=existing_committed_batches > 0,
        ),
        rows=rows,
        errors=errors,
    )


def commit_monthly_inventory(
    *,
    db: Session,
    payload: UploadCommitRequest,
) -> UploadBatch:
    existing_batches = list(
        db.scalars(
            select(UploadBatch).where(
                UploadBatch.year == payload.year,
                UploadBatch.month == payload.month,
                UploadBatch.source == payload.source,
                UploadBatch.status == UploadStatus.committed,
            )
        )
    )
    if existing_batches and not payload.replace_existing:
        raise ValueError("같은 year/month/source의 업로드가 이미 있습니다. replace_existing=true로 다시 요청하세요.")

    for batch in existing_batches:
        batch.status = UploadStatus.rolled_back

    upload_batch = UploadBatch(
        file_name=payload.file_name,
        year=payload.year,
        month=payload.month,
        source=payload.source,
        status=UploadStatus.committed,
        total_rows=len(payload.rows) + len(payload.errors),
        matched_rows=sum(1 for row in payload.rows if row.matched),
        unmatched_rows=sum(1 for row in payload.rows if not row.matched),
        error_rows=len(payload.errors),
    )
    db.add(upload_batch)
    db.flush()

    movement_date = date(payload.year, payload.month, monthrange(payload.year, payload.month)[1])

    products = list(db.scalars(select(Product)).all())

    for error in payload.errors:
        upload_batch.errors.append(
            _build_upload_error(upload_batch.id, error)
        )

    for row in payload.rows:
        product = _resolve_or_create_product(db=db, products=products, row=row)
        snapshot = InventorySnapshot(
            product_id=product.id,
            year=payload.year,
            month=payload.month,
            current_stock=row.current_stock,
            closing_stock_current_month=row.closing_stock_current_month,
            closing_stock_previous_month=row.closing_stock_previous_month,
            stock_change=row.stock_change,
            inbound_quantity=row.inbound_quantity,
            warehouse_inbound_quantity=row.warehouse_inbound_quantity,
            return_inbound_quantity=row.return_inbound_quantity,
            outbound_quantity=row.outbound_quantity,
            carryout_quantity=row.carryout_quantity,
            return_outbound_quantity=row.return_outbound_quantity,
            adjustment_inbound_quantity=row.adjustment_inbound_quantity,
            adjustment_outbound_quantity=row.adjustment_outbound_quantity,
            net_change=row.net_change,
            upload_batch_id=upload_batch.id,
        )
        db.add(snapshot)
        _add_stock_movements(
            db=db,
            product_id=product.id,
            movement_date=movement_date,
            upload_batch_id=upload_batch.id,
            row=row,
        )

    db.commit()
    db.refresh(upload_batch)
    return upload_batch


def list_upload_batches(db: Session) -> UploadBatchListResponse:
    batches = list(
        db.scalars(select(UploadBatch).order_by(UploadBatch.uploaded_at.desc(), UploadBatch.id.desc()))
    )
    return UploadBatchListResponse(
        items=[
            UploadBatchItem(
                id=batch.id,
                file_name=batch.file_name,
                year=batch.year,
                month=batch.month,
                source=batch.source,
                uploaded_at=batch.uploaded_at.isoformat(),
                status=batch.status.value,
                total_rows=batch.total_rows,
                matched_rows=batch.matched_rows,
                unmatched_rows=batch.unmatched_rows,
                error_rows=batch.error_rows,
            )
            for batch in batches
        ]
    )


def get_current_inventory(
    *,
    db: Session,
    search: str | None = None,
    low_stock_only: bool = False,
    year: int | None = None,
    month: int | None = None,
) -> InventoryListResponse:
    committed_batches = list(
        db.scalars(
            select(UploadBatch).where(UploadBatch.status == UploadStatus.committed)
        )
    )
    batch_ids = {batch.id for batch in committed_batches}
    if not batch_ids:
        return InventoryListResponse(items=[])

    snapshots = list(
        db.scalars(
            select(InventorySnapshot).where(InventorySnapshot.upload_batch_id.in_(batch_ids))
        )
    )
    latest_by_product: dict[int, InventorySnapshot] = {}
    for snapshot in snapshots:
        if year is not None and month is not None:
            if snapshot.year != year or snapshot.month != month:
                continue
        current = latest_by_product.get(snapshot.product_id)
        if current is None or (snapshot.year, snapshot.month, snapshot.id) > (
            current.year,
            current.month,
            current.id,
        ):
            latest_by_product[snapshot.product_id] = snapshot

    products = {
        product.id: product for product in db.scalars(select(Product)).all()
    }
    query = _normalize_name(search) if search else None
    items: list[InventoryListItem] = []
    for product_id, snapshot in latest_by_product.items():
        product = products.get(product_id)
        if not product:
            continue
        is_low_stock = snapshot.current_stock <= 0
        if query and query not in _normalize_name(
            " ".join(filter(None, [product.product_code, product.external_code, product.barcode, product.name]))
        ):
            continue
        if low_stock_only and not is_low_stock:
            continue
        items.append(
            InventoryListItem(
                product_id=product.id,
                product_code=product.product_code,
                external_code=product.external_code,
                barcode=product.barcode,
                name=product.name,
                category=product.category,
                current_stock=snapshot.current_stock,
                closing_stock_current_month=snapshot.closing_stock_current_month,
                closing_stock_previous_month=snapshot.closing_stock_previous_month,
                inbound_quantity=snapshot.inbound_quantity,
                outbound_quantity=snapshot.outbound_quantity,
                net_change=snapshot.net_change,
                snapshot_year=snapshot.year,
                snapshot_month=snapshot.month,
                is_low_stock=is_low_stock,
                latest_upload_batch_id=snapshot.upload_batch_id,
            )
        )

    items.sort(key=lambda item: (item.snapshot_year, item.snapshot_month, item.name), reverse=True)
    return InventoryListResponse(items=items)


def get_monthly_inventory(db: Session, *, year: int, month: int) -> InventoryMonthlyResponse:
    committed_batch_ids = {
        batch.id
        for batch in db.scalars(
            select(UploadBatch).where(UploadBatch.status == UploadStatus.committed)
        )
    }
    snapshots = list(
        db.scalars(
            select(InventorySnapshot).where(
                InventorySnapshot.year == year,
                InventorySnapshot.month == month,
                InventorySnapshot.upload_batch_id.in_(committed_batch_ids),
            )
        )
    )
    products = {product.id: product for product in db.scalars(select(Product)).all()}
    items = [
        InventoryMonthlyItem(
            product_id=snapshot.product_id,
            product_code=products[snapshot.product_id].product_code if snapshot.product_id in products else None,
            external_code=products[snapshot.product_id].external_code if snapshot.product_id in products else None,
            barcode=products[snapshot.product_id].barcode if snapshot.product_id in products else None,
            name=products[snapshot.product_id].name if snapshot.product_id in products else "",
            current_stock=snapshot.current_stock,
            closing_stock_current_month=snapshot.closing_stock_current_month,
            closing_stock_previous_month=snapshot.closing_stock_previous_month,
            stock_change=snapshot.stock_change,
            inbound_quantity=snapshot.inbound_quantity,
            warehouse_inbound_quantity=snapshot.warehouse_inbound_quantity,
            return_inbound_quantity=snapshot.return_inbound_quantity,
            outbound_quantity=snapshot.outbound_quantity,
            carryout_quantity=snapshot.carryout_quantity,
            return_outbound_quantity=snapshot.return_outbound_quantity,
            adjustment_inbound_quantity=snapshot.adjustment_inbound_quantity,
            adjustment_outbound_quantity=snapshot.adjustment_outbound_quantity,
            net_change=snapshot.net_change,
            upload_batch_id=snapshot.upload_batch_id,
        )
        for snapshot in snapshots
    ]
    items.sort(key=lambda item: item.name)
    return InventoryMonthlyResponse(year=year, month=month, items=items)


def split_external_product_name(value: str) -> tuple[str | None, str]:
    value = value.strip()
    if " / " not in value:
        return None, value
    external_code, product_name = value.split(" / ", 1)
    return external_code.strip() or None, product_name.strip()


def match_product(
    *,
    products: list[Product],
    product_code: str | None,
    barcode: str | None,
    product_name: str,
) -> ProductMatch:
    if product_code:
        for product in products:
            if _same(product.product_code, product_code):
                return ProductMatch(product=product, method="product_code")

    if barcode:
        for product in products:
            if _same(product.barcode, barcode):
                return ProductMatch(product=product, method="barcode")

    normalized_name = _normalize_name(product_name)
    best_product: Product | None = None
    best_ratio = 0.0
    for product in products:
        ratio = SequenceMatcher(None, normalized_name, _normalize_name(product.name)).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_product = product

    if best_product and best_ratio >= 0.72:
        return ProductMatch(product=best_product, method="name_similarity")

    return ProductMatch(product=None, method=None)


def _read_headers(sheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    first_row = next(sheet.iter_rows(min_row=1, max_row=1), [])
    for index, cell in enumerate(first_row, start=1):
        value = _text(cell.value)
        if value:
            headers[value] = index
    return headers


def _iter_data_rows(sheet, headers: dict[str, int]):
    for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        values = {}
        for column_name, column_index in headers.items():
            cell = row[column_index - 1] if len(row) >= column_index else None
            values[column_name] = None if cell is None else cell.value
        if any(_text(value) for value in values.values()):
            yield row_number, values


def _count_existing_batches(db: Session, year: int, month: int, source: str) -> int:
    return len(
        list(
            db.scalars(
                select(UploadBatch).where(
                    UploadBatch.year == year,
                    UploadBatch.month == month,
                    UploadBatch.source == source,
                    UploadBatch.status == UploadStatus.committed,
                )
            )
        )
    )


def _resolve_year_month(
    file_name: str, year: int | None, month: int | None
) -> tuple[int, int]:
    today = date.today()
    resolved_year = year or today.year
    resolved_month = month or today.month

    if year is None or month is None:
        match = re.search(r"(20\d{2})[-_. ]?(\d{1,2})[-_. ]?\d{1,2}", file_name)
        if match:
            resolved_year = year or int(match.group(1))
            resolved_month = month or int(match.group(2))

    if not 1 <= resolved_month <= 12:
        raise ValueError("month는 1~12 사이여야 합니다.")
    return resolved_year, resolved_month


def _duplicate_key(
    product_code: str | None,
    barcode: str | None,
    external_code: str | None,
    product_name: str,
) -> str:
    return "|".join(
        [
            product_code or "",
            barcode or "",
            external_code or "",
            _normalize_name(product_name),
        ]
    )


def _normalize_name(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"[^0-9a-zA-Z가-힣]+", "", value).lower()


def _same(left: str | None, right: str | None) -> bool:
    return _text(left) == _text(right)


def _text(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _clean_cell(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _to_int(value) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return 0
    return int(float(text))


def _resolve_or_create_product(
    *,
    db: Session,
    products: list[Product],
    row: UploadPreviewRow,
) -> Product:
    if row.matched and row.matched_product:
        for product in products:
            if product.id == row.matched_product.id:
                _hydrate_product(product, row)
                return product

    product = Product(
        product_code=row.product_code,
        external_code=row.external_code,
        barcode=row.barcode,
        name=row.product_name,
    )
    db.add(product)
    db.flush()
    products.append(product)
    return product


def _hydrate_product(product: Product, row: UploadPreviewRow) -> None:
    if row.product_code and row.product_code != product.product_code:
        product.code_histories.append(
            ProductCodeHistory(
                previous_product_code=product.product_code,
                new_product_code=row.product_code,
                changed_reason="monthly inventory upload sync",
            )
        )
        product.product_code = row.product_code
    if row.external_code:
        product.external_code = row.external_code
    if row.barcode:
        product.barcode = row.barcode
    if row.product_name:
        product.name = row.product_name


def _build_upload_error(upload_batch_id: int, error: UploadPreviewError):
    from app.models.upload import UploadError

    return UploadError(
        upload_batch_id=upload_batch_id,
        row_number=error.row_number,
        error_message=error.error_message,
        raw_data=error.raw_data,
    )


def _add_stock_movements(
    *,
    db: Session,
    product_id: int,
    movement_date: date,
    upload_batch_id: int,
    row: UploadPreviewRow,
) -> None:
    movement_specs = [
        ("inbound", MovementType.inbound, row.inbound_quantity),
        ("warehouse_inbound", MovementType.inbound, row.warehouse_inbound_quantity),
        ("return_inbound", MovementType.inbound, row.return_inbound_quantity),
        ("adjustment_inbound", MovementType.inbound, row.adjustment_inbound_quantity),
        ("outbound", MovementType.outbound, row.outbound_quantity),
        ("carryout", MovementType.outbound, row.carryout_quantity),
        ("return_outbound", MovementType.outbound, row.return_outbound_quantity),
        ("adjustment_outbound", MovementType.outbound, row.adjustment_outbound_quantity),
    ]
    for source, movement_type, quantity in movement_specs:
        if quantity <= 0:
            continue
        db.add(
            StockMovement(
                product_id=product_id,
                movement_date=movement_date,
                movement_type=movement_type,
                quantity=quantity,
                source=source,
                upload_batch_id=upload_batch_id,
                memo="monthly upload commit",
            )
        )
