from __future__ import annotations

from pathlib import Path
import sys

from openpyxl import load_workbook
from sqlalchemy import select

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.db.base import Base  # noqa: E402
from app.db.session import SessionLocal, engine  # noqa: E402
from app.models.material import Material  # noqa: E402
from app.models.product import Product  # noqa: E402


def main(workbook_path: str) -> None:
    path = Path(workbook_path)
    if not path.exists():
        raise SystemExit(f"파일을 찾을 수 없습니다: {path}")

    workbook = load_workbook(path, data_only=True, read_only=False)
    Base.metadata.create_all(bind=engine)

    with SessionLocal() as db:
        _import_products(db, workbook["2026"])
        _import_materials(db, workbook["부자재 재고현황"])
        db.commit()

    print("legacy bootstrap complete")


def _import_products(db, sheet) -> None:
    existing_by_code = {
        product.product_code: product
        for product in db.scalars(select(Product)).all()
        if product.product_code
    }
    for row in sheet.iter_rows(min_row=3, values_only=True):
        name = _text(row[0])
        code = _text(row[1])
        if not name:
            continue
        product = existing_by_code.get(code) if code else None
        if product is None:
            product = Product(product_code=code, name=name)
            db.add(product)
            if code:
                existing_by_code[code] = product
        else:
            product.name = name


def _import_materials(db, sheet) -> None:
    existing_keys = {
        (material.name, material.material_type): material
        for material in db.scalars(select(Material)).all()
    }
    current_product_name = None
    for row in sheet.iter_rows(min_row=4, values_only=True):
        base_name = _text(row[0])
        material_type = _text(row[1])
        if base_name:
            current_product_name = base_name.replace("\n", " ").strip()
        if not current_product_name or not material_type:
            continue

        name = current_product_name
        current_stock = _to_int(row[19] if len(row) > 19 else 0)
        safety_stock = _to_int(row[17] if len(row) > 17 else 0)
        vendor = _text(row[23] if len(row) > 23 else None)
        unit_price = _to_int(row[24] if len(row) > 24 else 0)
        memo = _text(row[28] if len(row) > 28 else None)

        key = (name, material_type)
        material = existing_keys.get(key)
        if material is None:
            material = Material(
                name=name,
                material_type=material_type,
                current_stock=current_stock,
                safety_stock=safety_stock,
                vendor=vendor,
                unit_price=unit_price or None,
                memo=memo,
            )
            db.add(material)
            existing_keys[key] = material
        else:
            material.current_stock = current_stock
            material.safety_stock = safety_stock
            material.vendor = vendor
            material.unit_price = unit_price or None
            material.memo = memo


def _text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _to_int(value) -> int:
    if value in (None, ""):
        return 0
    return int(float(str(value).replace(",", "")))


if __name__ == "__main__":
    if len(sys.argv) != 2:
        raise SystemExit("usage: python scripts/bootstrap_legacy_inventory.py <xlsx_path>")
    main(sys.argv[1])
